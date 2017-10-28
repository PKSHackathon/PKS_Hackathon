import openpyxl
wb = openpyxl.load_workbook('baza.xlsx')
event = wb.get_sheet_by_name('Arkusz1')
i = 53
while i > 3:
    a = Event()
    if event.cell(row=i, column=2).value:
        a.nazwa = event.cell(row=i, column=2).value
    else:
        a.nazwa = ""
    if event.cell(row=i, column=3).value:
        a.data = event.cell(row=i, column=3).value
    else:
        a.data = ""
    if event.cell(row=i, column=4).value:
        a.godzina = event.cell(row=i, column=4).value
    else:
        a.godzina = ""
    if event.cell(row=i, column=5).value:
        a.miejsce = event.cell(row=i, column=5).value
    else:
        a.miejsce = ""
    if event.cell(row=i, column=6).value:
        a.uwagi = event.cell(row=i, column=6).value
    else:
        a.uwagi = ""
    a.save()
    i = i-1
