import openpyxl
events= [None]
y = 0
wb = openpyxl.load_workbook('baza.xlsx')                    # Załadowanie bazy
wb.get_sheet_names()                                        # Pobranie nazw arkuszy
event = wb.get_sheet_by_name('Arkusz1')                     # Przypisanie arkusza do zmiennej event
for i in range(4, 54, 1):                                   # Przejście przez wiersze
    for z in range(2, 7, 1):                                # Przejście przez komórki
        if event.cell(row=i, column=z).value:               # Sprawdzanie czy komórka zawiera dane
            print(event.cell(row=i, column=z).value)
    else:
        y += 1
    print("=============")
